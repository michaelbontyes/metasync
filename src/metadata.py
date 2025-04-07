import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os
import tempfile
import numpy as np
import time

DEFAULT_FILE = os.path.abspath("metadata-example.xlsx")
ENGINE = 'openpyxl'

def save_excel_with_sheet(df, file_path, sheet_name):
    """Save a single sheet while preserving other sheets"""
    try:
        if not isinstance(df, pd.DataFrame):
            raise ValueError("Expected DataFrame, got {}".format(type(df)))

        # Create temp file
        temp_path = tempfile.mktemp(suffix='.xlsx')
        
        # Load existing workbook
        book = load_workbook(file_path)
        
        # Create new writer with mode='a' for append
        with pd.ExcelWriter(temp_path, engine=ENGINE, mode='w') as writer:
            # Copy all sheets
            for sheet in book.sheetnames:
                if sheet == sheet_name:
                    # Write edited sheet
                    df.to_excel(writer, sheet_name=sheet, index=False)
                else:
                    # Copy existing sheet
                    sheet_df = pd.read_excel(file_path, sheet_name=sheet, engine=ENGINE)
                    sheet_df.to_excel(writer, sheet_name=sheet, index=False)
            
        # Replace original file
        os.replace(temp_path, file_path)
        return True
        
    except Exception as e:
        st.error(f"Save failed: {str(e)}")
        if os.path.exists(temp_path):
            os.unlink(temp_path)
        return False

def clean_data(df):
    """Handle mixed data types and NaN values"""
    for col in df.columns:
        # Convert mixed numeric columns
        if pd.api.types.is_string_dtype(df[col]):
            try:
                df[col] = pd.to_numeric(df[col], errors='ignore')
            except:
                pass
        # Fill NaN with empty string for string columns
        if pd.api.types.is_string_dtype(df[col]):
            df[col] = df[col].fillna('')
    return df

def save_changes(sheet, edited_df):
    """Save changes to Excel file and update state"""
    if not st.session_state.save_in_progress[sheet]:
        st.session_state.save_in_progress[sheet] = True
        
        with st.spinner(f"Saving changes to {sheet}..."):
            # First update state
            updated_df = apply_pending_changes(sheet, edited_df)
            
            # Then save to file
            if save_excel_with_sheet(updated_df, DEFAULT_FILE, sheet):
                st.session_state[f"previous_df_{sheet}"] = updated_df.copy()
                st.session_state.pending_changes[sheet] = {}
                st.session_state.last_change_saved[sheet] = True
                st.toast(f"Saved changes to {sheet}!", icon="✅")
                
        st.session_state.save_in_progress[sheet] = False
        return True
    return False

def main():
    st.set_page_config(layout="wide")
    st.title("Metadata Editor")
    
    # Initialize session state for tracking changes
    if "pending_changes" not in st.session_state:
        st.session_state.pending_changes = {}
    if "last_change_saved" not in st.session_state:
        st.session_state.last_change_saved = {}
    if "save_in_progress" not in st.session_state:
        st.session_state.save_in_progress = {}
    if "autosave_enabled" not in st.session_state:
        st.session_state.autosave_enabled = True
    
    # Autosave toggle
    st.sidebar.title("Settings")
    st.session_state.autosave_enabled = st.sidebar.checkbox(
        "Enable autosave", 
        value=st.session_state.autosave_enabled,
        help="Automatically save changes when data is modified"
    )
    
    if os.path.exists(DEFAULT_FILE):
        try:
            xls = pd.ExcelFile(DEFAULT_FILE, engine=ENGINE)
            sheets = xls.sheet_names
            
            # Create tabs for each sheet
            tabs = st.tabs(sheets)
            
            for i, sheet in enumerate(sheets):
                with tabs[i]:
                    # Initialize sheet-specific session state
                    if f"previous_df_{sheet}" not in st.session_state:
                        df = pd.read_excel(DEFAULT_FILE, sheet_name=sheet, engine=ENGINE)
                        df = clean_data(df)
                        st.session_state[f"previous_df_{sheet}"] = df.copy()
                        st.session_state.pending_changes[sheet] = {}
                        st.session_state.last_change_saved[sheet] = True
                        st.session_state.save_in_progress[sheet] = False
                    
                    # Load and clean sheet data
                    df = pd.read_excel(DEFAULT_FILE, sheet_name=sheet, engine=ENGINE)
                    df = clean_data(df)
                    
                    # Data editor with built-in filters
                    edited_df = st.data_editor(
                        df,
                        num_rows="dynamic",
                        key=f"editor_{sheet}",
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "_index": st.column_config.NumberColumn(disabled=True),
                        }
                    )
                    
                    # Detect changes when data editor is modified
                    if st.session_state.get(f"editor_{sheet}") is not None:
                        has_changes = detect_changes(sheet, st.session_state[f"previous_df_{sheet}"], edited_df)
                        
                        # Autosave on change detection
                        if has_changes and st.session_state.autosave_enabled and not st.session_state.last_change_saved[sheet]:
                            save_changes(sheet, edited_df)
                    
                    # Show pending changes indicator
                    if sheet in st.session_state.pending_changes and st.session_state.pending_changes[sheet]:
                        st.info(f"{len(st.session_state.pending_changes[sheet])} unsaved changes")
                    
                    # Manual save button
                    if st.button(f"Save {sheet}", key=f"save_{sheet}"):
                        save_changes(sheet, edited_df)
                        
        except Exception as e:
            st.error(f"Error: {str(e)}")
    else:
        st.error(f"File not found: {DEFAULT_FILE}")

def detect_changes(sheet, previous_df, current_df):
    """Detect changes between previous and current dataframes"""
    changes = {}
    has_new_changes = False
    
    # Compare dataframes cell by cell
    for col in previous_df.columns:
        for idx in previous_df.index:
            if idx < len(current_df) and col in current_df.columns:
                # Check if values are different
                prev_val = previous_df.at[idx, col]
                curr_val = current_df.at[idx, col]
                
                if pd.isna(prev_val) and pd.isna(curr_val):
                    continue
                elif pd.isna(prev_val) and not pd.isna(curr_val):
                    changes[(idx, col)] = curr_val
                    has_new_changes = True
                elif not pd.isna(prev_val) and pd.isna(curr_val):
                    changes[(idx, col)] = None
                    has_new_changes = True
                elif prev_val != curr_val:
                    changes[(idx, col)] = curr_val
                    has_new_changes = True
    
    # Store changes in session state
    st.session_state.pending_changes[sheet] = changes
    
    # Mark that we have unsaved changes
    if has_new_changes:
        st.session_state.last_change_saved[sheet] = False
        
    return has_new_changes

def apply_pending_changes(sheet, df):
    """Apply pending changes to the dataframe"""
    # Create a copy to avoid modifying the original
    updated_df = df.copy()
    
    # Apply each pending change
    for (idx, col), value in st.session_state.pending_changes[sheet].items():
        if idx < len(updated_df) and col in updated_df.columns:
            updated_df.at[idx, col] = value
    
    return updated_df

if __name__ == "__main__":
    main()
